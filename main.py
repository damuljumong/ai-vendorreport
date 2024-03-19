import streamlit as st
from langchain.chat_models import ChatOpenAI
#from langchain.llms import CTransformers
import os
import OpenDartReader

import pandas as pd
from streamlit_extras.buy_me_a_coffee import button
import yfinance as yf
from datetime import datetime, timedelta

import openpyxl
import tempfile
#import docx

def main():

    # while True:  # 무한 루프 시작
        st.set_page_config(page_title='VendorReport', page_icon = 'buybeer_30.png', initial_sidebar_state = 'auto')
        # HTML 코드를 직접 추가
        adhtml_code = """
        <ins class="kakao_ad_area" style="display:none;"
        data-ad-unit="DAN-BVSIspEij6yfBT9F"
        data-ad-width="320"
        data-ad-height="100"></ins>
        <script type="text/javascript" src="//t1.daumcdn.net/kas/static/ba.min.js" async></script>
        """
        # streamlit 앱에 HTML 코드 추가
        st.write(adhtml_code, unsafe_allow_html=True)
    
        # button(username="damuljumong", floating=True, width=221)
        # Buy Me a Coffee HTML 코드
        tossme_button = """        
        <div style="position: absolute; bottom: 10px; right: 10px;">
            <a href="https://toss.me/damulcandy" target="_blank">
                <img src="https://harlequin-national-unicorn-728.mypinata.cloud/ipfs/QmNj9VLSE1GpoP4sS9q8TMwbRh6FS9XWTTFNx8hUr4L9AW/buybeer_50.png" alt="insert coin" style="height: 50px !important;width: 50px !important;" >
            </a>
        </div>
        """
        
        # Streamlit 앱에 HTML 삽입
        st.write(tossme_button, unsafe_allow_html=True)
    
        chat_model = ChatOpenAI()

        #api_key = 'API_KEY_DART'
        api_key = st.secrets["API_KEY_DART"]
        dart = OpenDartReader(api_key) 

        st.title('국내 업체 경영현황 보고서')

        stock_codes_input = "005930,072130,078000,069410" # 삼성전자 , 유엔젤, 텔코웨어 엔텔스 
        symbol = "005930"
    
        end_date = datetime.now()
        start_date = end_date - timedelta(days=365 * 5)  # 5년간의 데이터

        year = 2023
        financial_records = []
        vendorinfo_records = []
    
        # Streamlit title
        st.title("조회할 모든 업체 리스트(STOCK CODE)를 엑셀문서 xlsx 표준 포맷으로 한번에 모든 업체를 조회하세요")
        st.write("---")
        # File upload
        uploaded_file = st.file_uploader(" .xlsx 형식의 엑셀 문서 파일을 올려주세요!", type=['xlsx'])

        stock_codes_input = ""  # 초기값 설정
        
        if uploaded_file is not None:
            # Temporary directory for file processing
            temp_dir = tempfile.TemporaryDirectory()
            temp_filepath = os.path.join(temp_dir.name, uploaded_file.name)
            with open(temp_filepath, "wb") as f:
                f.write(uploaded_file.getvalue())
        
            # Determine the file type and process accordingly
            file_extension = os.path.splitext(uploaded_file.name)[-1].lower()
            if file_extension == ".xlsx":
                # Process xlsx file, 첫 번째 시트의 첫 번째 열만 처리
                wb = openpyxl.load_workbook(temp_filepath)
                sheet = wb[wb.sheetnames[0]]  # 첫 번째 시트
                for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    # 첫 번째 행을 제외하고 처리
                    if row_index > 1 and row[0]:  # 2번째 행부터, 첫 번째 열이 비어 있지 않은 경우에만 추가
                        stock_codes_input += str(row[0]) + ","
        
                stock_codes_input = stock_codes_input.rstrip(",")  # 마지막 쉼표 제거
        
        # 사용자 입력 필드에 읽어들인 주식 코드 표시
        user_stock_codes_input = st.text_input('업체 Stock code를 입력하세요. 예: 삼성전자 Stock code 005930', value=stock_codes_input, key='stock_codes_input_1')
        
        if user_stock_codes_input:
            stock_codes = [code.strip() for code in user_stock_codes_input.split(',')]
    
                
        #stock_codes_input = st.text_input('업체 Stock code를 입력하세요. 예 삼성전자 Stock code 005930,072130,078000,069410',key='stock_codes_input_1')
        #if stock_codes_input:
        #    stock_codes = [code.strip() for code in stock_codes_input.split(',')]

        #content = st.text_input('인공지능이 분석할 업체명을 입력하세요.')

        if st.button('업체 분석 요청'):
            with st.spinner('업체 리포트 작성 중...'):
                for symbol in stock_codes:
                    # 회사명에  포함된 회사들에 대한 개황정보
                    vendorInfo = dart.company(symbol)
                    # excel_filename = f'./files/{symbol}_vendorInfo.xlsx'
                    # vendorInfo.to_excel(excel_filename, index=False)             
                    PeopleInfo = dart.report(symbol, '직원', year - 1)
                    # excel_filename = f'./files/{symbol}_PeopleInfo.xlsx'
                    # PeopleInfo.to_excel(excel_filename, index=False)     

                    fnInfo = dart.finstate(symbol, year -1,reprt_code ='11011') 
                    # excel_filename = f'./files/{symbol}_fnInfo.xlsx'
                    # fnInfo.to_excel(excel_filename, index=False) 
                    fnInfo_1Q = dart.finstate(symbol, year,reprt_code ='11013') # 1 분기
                    # excel_filename = f'./files/{symbol}_fnInfo_1Q.xlsx'
                    # fnInfo_1Q.to_excel(excel_filename, index=False)  
                    fnInfo_2Q = dart.finstate(symbol, year,reprt_code ='11012') # 2 분기
                    # excel_filename = f'./files/{symbol}_fnInfo_2Q.xlsx'
                    # fnInfo_2Q.to_excel(excel_filename, index=False)  
                    fnInfo_3Q = dart.finstate(symbol, year,reprt_code ='11014') # 3 분기
                    # excel_filename = f'./files/{symbol}_fnInfo_3Q.xlsx'
                    # fnInfo_3Q.to_excel(excel_filename, index=False)

                    # 선택할 행과 열 인덱스
                    # selected_rows = [0,1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]

                    # # fnInfo 데이터프레임에서 선택된 행과 열 선택
                    # selected_vendorInfo = vendorInfo.iloc[selected_rows, [4, 9, 18, 15, 12]]
                    # selected_vendorInfo.columns = ['회사명', '회계항목', '20년', '21년', '22년']  # Rename columns

                    # # fnInfo_1Q, fnInfo_2Q, fnInfo_3Q에서 선택된 행과 열 선택
                    # selected_PeopleInfo = PeopleInfo.iloc[selected_rows, [12]]
                    # selected_PeopleInfo.columns = ['23년 1Q']  # Rename columns

                    previous_corp_name = None
                    previous_ceo_nm = None
                    previous_adres = None
                    previous_est_dt = None                

                    # sm, fo_bbm, sexdstn 값을 리스트로 저장
                    fo_bbm_list = PeopleInfo['fo_bbm']                    
                    sexdstn_list = PeopleInfo['sexdstn']
                    sm_list = PeopleInfo['sm']
                    for fo_bbm, sexdstn, sm in zip(                    
                        fo_bbm_list, 
                        sexdstn_list,
                        sm_list         
                    ):
                        corp_name = vendorInfo['corp_name']
                        ceo_nm = vendorInfo['ceo_nm']
                        adres = vendorInfo['adres']
                        est_dt = vendorInfo['est_dt']
                        fo_bbm_list = PeopleInfo['fo_bbm']                    
                        sexdstn_list = PeopleInfo['sexdstn']
                        sm_list = PeopleInfo['sm']
                        # corp_name이 이전과 동일한 경우에만 추가
                        if ( corp_name != previous_corp_name or ceo_nm != previous_ceo_nm or adres != previous_adres or est_dt != previous_est_dt ):
                            vendorinfo_records.append({
                                'corp_name': corp_name,
                                'ceo_nm': ceo_nm,
                                'adres': adres,
                                'est_dt': est_dt,
                                'Business': fo_bbm,
                                'sex': sexdstn,
                                'employees': sm     
                            })
                        else :
                            vendorinfo_records.append({
                                'Business': fo_bbm,
                                'sex': sexdstn,
                                'employees': sm     
                            })
                        # 이전 corp_name 업데이트
                        previous_corp_name = corp_name
                        previous_ceo_nm = ceo_nm
                        previous_adres = adres
                        previous_est_dt = est_dt                


                    # st.dataframe(vendorInfo,width=800)
                    # st.dataframe(PeopleInfo,width=800)
                    # 기존의 리스트를 데이터프레임으로 변환
                    vendorinfo_df = pd.DataFrame(vendorinfo_records)
                    # 열 이름을 변경
                    vendorinfo_df.columns = ['회사명', '대표', '주소', '설립일', '사업부', '성별', '종업원수']        
                    st.dataframe(vendorinfo_df,width=800)
                    vendorinfo_records = []

                    # 선택할 행과 열 인덱스
                    selected_rows = [0,1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]

                    # fnInfo 데이터프레임에서 선택된 행과 열 선택
                    selected_fnInfo = fnInfo.iloc[selected_rows, [4, 9, 18, 15, 12]]
                    selected_fnInfo.columns = ['회사명', '회계항목', '20년', '21년', '22년']  # Rename columns

                    # fnInfo_1Q, fnInfo_2Q, fnInfo_3Q에서 선택된 행과 열 선택
                    selected_fnInfo_1Q = fnInfo_1Q.iloc[selected_rows, [12]]
                    selected_fnInfo_1Q.columns = ['23년 1Q']  # Rename columns

                    selected_fnInfo_2Q = fnInfo_2Q.iloc[selected_rows, [12]]
                    selected_fnInfo_2Q.columns = ['23년 2Q']  # Rename columns

                    # fnInfo_3Q 데이터프레임을 위한 초기화
                    selected_fnInfo_3Q = None

                    # fnInfo_3Q가 비어있지 않은 경우에만 처리
                    if not fnInfo_3Q.empty:
                        # 선택한 행의 인덱스 중에서 데이터프레임의 길이를 초과하는 인덱스를 필터링하여 출력
                        selected_rows = [index for index in selected_rows if index < len(fnInfo_3Q)]

                        # 선택한 행이 존재하는 경우에만 fnInfo_3Q에서 선택된 열 선택
                        if selected_rows:
                            selected_fnInfo_3Q = fnInfo_3Q.iloc[selected_rows, [12]]
                            selected_fnInfo_3Q.columns = ['23년 3Q']  # Rename columns

                    # 선택된 데이터프레임을 하나로 통합
                    combined_df = pd.concat([selected_fnInfo, selected_fnInfo_1Q, selected_fnInfo_2Q, selected_fnInfo_3Q], axis=1)

                    # 데이터프레임 출력
                    st.dataframe(combined_df, width=1200)

                    # Event 
                    # dart.event(corp, event, start=None, end=None)
                    # 조회가능한 주요사항 항목: 
                    # ['부도발생', '영업정지', '회생절차', '해산사유', '유상증자', '무상증자', '유무상증자', '감자', '관리절차개시', '소송', '해외상장결정', '해외상장폐지결정', '해외상장', '해외상장폐지', '전환사채발행', '신주인수권부사채발행', '교환사채발행', '관리절차중단', '조건부자본증권발행', '자산양수도', '타법인증권양도', '유형자산양도', '유형자산양수', '타법인증권양수', '영업양도', '영업양수', '자기주식취득신탁계약해지', '자기주식취득신탁계약체결', '자기주식처분', '자기주식취득', '주식교환', '회사분할합병', '회사분할', '회사합병', '사채권양수', '사채권양도결정']
                    vn_event =  ['부도발생', '영업정지', '회생절차', '해산사유', '유상증자', '무상증자', '유무상증자', '감자', '관리절차개시', '소송', '해외상장결정', '해외상장폐지결정', '해외상장', '해외상장폐지', '전환사채발행', '신주인수권부사채발행', '교환사채발행', '관리절차중단', '조건부자본증권발행', '자산양수도', '타법인증권양도', '유형자산양도', '유형자산양수', '타법인증권양수', '영업양도', '영업양수', '자기주식취득신탁계약해지', '자기주식취득신탁계약체결', '자기주식처분', '자기주식취득', '주식교환', '회사분할합병', '회사분할', '회사합병', '사채권양수', '사채권양도결정']
                    for each_event in vn_event:
                        issues = dart.event(symbol, each_event) #
                        if not issues.empty:
                            st.write(each_event)
                            st.dataframe(issues, width=1200)
                    #stock = 'AAPL'
                    #start_date = '2020-01-01'
                    #end_date = '2021-01-01'
                    #data = yf.download(symbol, start=start_date, end=end_date) 
                    #data =https://api.finance.naver.com/siseJson.naver?symbol=005930&requestType=1&startTime=20190624&endTime=20240315&timeframe=day
                    #data = fdr.DataReader(symbol,start=start_date, end=end_date)
                    #st.dataframe(data, width=1200)
            #st.write("인공지능 ( Open AI )이 분석한 기업 정보를 알려드립니다")
            #result = chat_model.predict(content + "을 분석해줘")    # OpenAI sknam
            #st.write(result)
            # HTML 코드를 직접 추가
            # 광고 코드 중복 추가 방지
            if not st.session_state.get("ad_added", False):
                adhtml_code2 = """
                <ins class="kakao_ad_area" style="display:none;"
                data-ad-unit="DAN-BVSIspEij6yfBT9F"
                data-ad-width="320"
                data-ad-height="100"></ins>
                <script type="text/javascript" src="//t1.daumcdn.net/kas/static/ba.min.js" async></script>
                """
                
                # streamlit 앱에 HTML 코드 추가
                st.write(adhtml_code2, unsafe_allow_html=True)
        # 광고가 이미 추가되었음을 표시
                st.session_state.ad_added = True
       
if __name__ == '__main__':
    main()
